"""
AI Universal Import — multi-format file analysis with OpenAI classification
and admin-approved sync into wells / locators / ro_trains and matching
*_readings tables. Every approve/reject decision is recorded in the
deletion_audit_log with an [IMPORT] tag.

Pipeline:
    1. extract_tables(filename, bytes) -> list[ExtractedTable]
    2. classify_tables(tables) -> list[TableAnalysis]   (OpenAI, with rule-based fallback)
    3. persist analysis row in `import_analysis`
    4. sync_analysis(decisions) -> creates entities, inserts readings,
       updates analysis status, writes [IMPORT] audit rows.

Design notes:
    * Wellmeter tri-block files are detected up-front and flagged so the UI
      can hand off to the proven legacy parser (`/api/import/parse-wellmeter`).
    * Table extraction is bounded (rows/tables/sample-bytes) so a malicious
      upload can't OOM the worker.
    * The OpenAI key is read from `OPENAI_API_KEY`; if missing, a rule-based
      header-keyword classifier is used so the workflow never hard-fails.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Optional

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from supabase import Client

# Auth + audit helpers (public names live in admin_helpers; admin_service
# re-exports them for legacy callers).
from admin_helpers import (
    AuditEntry,
    bearer_token,
    caller_identity,
    require_roles,
    user_scoped_client,
    write_audit,
)
from ai_import_helpers import (
    AnalysisPersistPayload,
    AuditDecision,
    ReadingsInsertContext,
    build_reading_rows,
    ensure_entity,
    to_jsonable,
)

log = logging.getLogger(__name__)

# ---- Tunables -------------------------------------------------------------
OPENAI_MODEL = os.environ.get("OPENAI_IMPORT_MODEL", "gpt-4o-mini")
MAX_TABLES = 12
MAX_BODY_ROWS = 500          # rows kept per table (after header detection)
MAX_SAMPLE_ROWS = 25         # rows shown to LLM and persisted as preview
MAX_CELL_LEN = 80            # truncation for sample cell text
MAX_FILE_BYTES = 25 * 1024 * 1024  # 25 MiB

SUPPORTED_TARGETS: set[str] = {
    "wells", "locators", "ro_trains",
    "well_readings", "locator_readings", "ro_train_readings", "power_readings",
    "skip", "unknown",
}

ENTITY_TARGETS = {"wells", "locators", "ro_trains"}
READING_TARGETS = {
    "well_readings", "locator_readings", "ro_train_readings", "power_readings",
}

# ---------------------------------------------------------------------------
# Multi-format extraction
# ---------------------------------------------------------------------------

@dataclass
class ExtractedTable:
    source: str
    headers: list[str]
    rows: list[list[Any]]          # full body (capped at MAX_BODY_ROWS)
    sample_csv: str                # CSV string sent to the LLM


def extract_tables(filename: str, content: bytes) -> list[ExtractedTable]:
    if len(content) > MAX_FILE_BYTES:
        raise HTTPException(status_code=413, detail="File too large (limit 25 MiB).")
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _extract_xlsx(content)
    if name.endswith(".docx"):
        return _extract_docx(content)
    if name.endswith(".doc"):
        raise HTTPException(
            status_code=400,
            detail="Legacy .doc binaries are not supported — please save as .docx and retry.",
        )
    if name.endswith((".txt", ".csv", ".tsv")):
        return _extract_text(content, filename or "upload.txt")
    raise HTTPException(
        status_code=400,
        detail=f"Unsupported file type: {filename}. Allowed: .xlsx, .xlsm, .docx, .txt, .csv, .tsv",
    )


def _extract_xlsx(content: bytes) -> list[ExtractedTable]:
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not open spreadsheet: {e}")
    out: list[ExtractedTable] = []
    for sheet in wb.sheetnames[:MAX_TABLES]:
        ws = wb[sheet]
        rows: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_BODY_ROWS + 5:
                break
            rows.append(list(row))
        headers, body = _pick_headers(rows)
        if not headers:
            continue
        body = body[:MAX_BODY_ROWS]
        out.append(ExtractedTable(
            source=sheet, headers=headers, rows=body,
            sample_csv=_to_csv(headers, body[:MAX_SAMPLE_ROWS]),
        ))
    return out


def _extract_docx(content: bytes) -> list[ExtractedTable]:
    try:
        from docx import Document  # type: ignore
    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"python-docx not installed: {e}",
        )
    try:
        doc = Document(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not open .docx: {e}")
    out: list[ExtractedTable] = []
    for ti, t in enumerate(doc.tables[:MAX_TABLES]):
        rows = [[cell.text for cell in r.cells] for r in t.rows[: MAX_BODY_ROWS + 5]]
        headers, body = _pick_headers(rows)
        if not headers:
            continue
        body = body[:MAX_BODY_ROWS]
        out.append(ExtractedTable(
            source=f"Table {ti + 1}", headers=headers, rows=body,
            sample_csv=_to_csv(headers, body[:MAX_SAMPLE_ROWS]),
        ))
    return out


def _extract_text(content: bytes, filename: str) -> list[ExtractedTable]:
    text = content.decode("utf-8", errors="replace")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = []
    for i, r in enumerate(reader):
        if i >= MAX_BODY_ROWS + 5:
            break
        rows.append(r)
    headers, body = _pick_headers(rows)
    if not headers:
        return []
    body = body[:MAX_BODY_ROWS]
    return [ExtractedTable(
        source=filename, headers=headers, rows=body,
        sample_csv=_to_csv(headers, body[:MAX_SAMPLE_ROWS]),
    )]


def _pick_headers(rows: list[list[Any]]) -> tuple[list[str], list[list[Any]]]:
    """Find the first row that looks like a header (>=2 non-empty,
    non-numeric cells) and treat everything after it as the body."""
    for i, r in enumerate(rows):
        nonempty = [c for c in r if c not in (None, "")]
        if len(nonempty) < 2:
            continue
        non_numeric = [c for c in nonempty if not _looks_numeric(c)]
        if len(non_numeric) >= max(2, int(len(nonempty) * 0.5)):
            return [_clean_header(c) for c in r], rows[i + 1 :]
    if rows:
        return [f"col{i + 1}" for i in range(len(rows[0]))], rows
    return [], []


def _clean_header(c: Any) -> str:
    if c is None:
        return ""
    s = str(c).strip()
    return re.sub(r"\s+", " ", s)


def _looks_numeric(v: Any) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).replace(",", ""))
        return True
    except (TypeError, ValueError):
        return False


def _to_csv(headers: list[str], rows: list[list[Any]]) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([_truncate(h) for h in headers])
    for r in rows:
        w.writerow([_truncate(c) for c in r])
    return buf.getvalue()


def _truncate(c: Any) -> str:
    if c is None:
        return ""
    s = str(c)
    return s if len(s) <= MAX_CELL_LEN else s[: MAX_CELL_LEN - 1] + "…"


# ---------------------------------------------------------------------------
# AI classification (OpenAI; rule-based fallback)
# ---------------------------------------------------------------------------

CLASSIFY_SYSTEM = """You are a data-import classifier for PWRI Monitoring (a multi-plant water-treatment app).
Given a single table extracted from an uploaded file, decide which application module it belongs to.

Available targets:
  wells              : Well master data (creating new wells: name, status).
  locators           : Locator (delivery point) master data.
  ro_trains          : RO train master data.
  well_readings      : Daily well meter readings (date + initial/final/volume).
  locator_readings   : Locator delivery readings (date + volume).
  ro_train_readings  : RO train production readings (date + permeate/feed flow).
  power_readings     : Power meter readings (date + kWh).
  skip               : Summary / header / footer table with no usable data.
  unknown            : No target fits — admin must decide.

For every table return a single JSON object (no prose):
{
  "target": "<one of above>",
  "entity_name": "<best guess for the well/locator/train this table is about, or null>",
  "confidence": 0.0-1.0,
  "column_mapping": { "<our_field>": "<source_header>" },
  "anomalies": ["short notes on data quality issues"],
  "rationale": "one sentence why"
}

Rules:
  * Be conservative: pick "unknown" rather than guess.
  * Confidence < 0.5 means the admin should review carefully.
  * For *_readings, column_mapping should include at minimum "date" (and "volume" or "value").
  * For wells/locators/ro_trains, column_mapping should include "name" (and "address" if present).
  * Flag obvious issues (missing dates, all-zero readings, header repetition) in `anomalies`.
"""


@dataclass
class TableAnalysis:
    source: str
    headers: list[str]
    target: str
    entity_name: Optional[str]
    confidence: float
    column_mapping: dict[str, str]
    anomalies: list[str]
    rationale: str
    sample_rows: list[list[str]]  # truncated string cells, safe for JSON
    row_count: int


def classify_tables(tables: list[ExtractedTable]) -> tuple[list[TableAnalysis], str]:
    """Returns (analyses, provider_label)."""
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return [_heuristic_classify(t) for t in tables], "rule-based"
    try:
        from openai import OpenAI  # type: ignore
        client = OpenAI(api_key=api_key)
    except Exception as e:
        log.warning("OpenAI init failed (%s); using heuristics", e)
        return [_heuristic_classify(t) for t in tables], "rule-based"

    out: list[TableAnalysis] = []
    for t in tables:
        try:
            resp = client.chat.completions.create(
                model=OPENAI_MODEL,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": CLASSIFY_SYSTEM},
                    {"role": "user", "content":
                        f"Source: {t.source}\nHeaders: {t.headers}\n\nSample rows (CSV):\n{t.sample_csv}"},
                ],
                temperature=0.1,
                timeout=30,
            )
            data = json.loads(resp.choices[0].message.content or "{}")
            target = str(data.get("target") or "unknown")
            if target not in SUPPORTED_TARGETS:
                target = "unknown"
            out.append(TableAnalysis(
                source=t.source,
                headers=t.headers,
                target=target,
                entity_name=(str(data["entity_name"]).strip() if data.get("entity_name") else None) or None,
                confidence=max(0.0, min(1.0, float(data.get("confidence") or 0.0))),
                column_mapping={str(k): str(v) for k, v in (data.get("column_mapping") or {}).items()},
                anomalies=[str(x) for x in (data.get("anomalies") or [])][:10],
                rationale=str(data.get("rationale") or "")[:300],
                sample_rows=[[_truncate(c) for c in r] for r in t.rows[:MAX_SAMPLE_ROWS]],
                row_count=len(t.rows),
            ))
        except Exception as e:
            log.exception("AI classify failed for %s", t.source)
            ta = _heuristic_classify(t)
            ta.anomalies.insert(0, f"AI call failed, used heuristics: {e}")
            out.append(ta)
    return out, "openai"


def _heuristic_classify(t: ExtractedTable) -> TableAnalysis:
    h = " ".join(str(x) for x in t.headers).lower()
    target = "unknown"
    if any(k in h for k in ("kwh", "power", "electric")):
        target = "power_readings"
    elif "ro" in h and any(k in h for k in ("permeate", "feed flow", "train")):
        target = "ro_train_readings"
    elif any(k in h for k in ("locator", "delivery point")):
        target = "locator_readings"
    elif any(k in h for k in ("initial", "final", "volume")) and any(k in h for k in ("date", "day")):
        target = "well_readings"
    elif "well" in h and any(k in h for k in ("status", "address", "type", "name")):
        target = "wells"
    return TableAnalysis(
        source=t.source,
        headers=t.headers,
        target=target,
        entity_name=t.source,
        confidence=0.4 if target != "unknown" else 0.1,
        column_mapping={},
        anomalies=["Classified by header keywords only (no AI key configured)."],
        rationale="Rule-based fallback (no LLM).",
        sample_rows=[[_truncate(c) for c in r] for r in t.rows[:MAX_SAMPLE_ROWS]],
        row_count=len(t.rows),
    )


# ---------------------------------------------------------------------------
# Wellmeter signature
# ---------------------------------------------------------------------------

def looks_like_wellmeter(tables: list[ExtractedTable]) -> bool:
    """Tri-block monthly layout: many `Initial`/`Final` columns repeated."""
    for t in tables:
        h = " ".join(str(x).lower() for x in t.headers)
        if h.count("initial") >= 2 and h.count("final") >= 2:
            return True
    return False


# ---------------------------------------------------------------------------
# Persistence
# ---------------------------------------------------------------------------

def _public_table_payload(a: TableAnalysis) -> dict[str, Any]:
    """Wire-format returned to the UI — preview-only, no body_rows."""
    return {
        "source": a.source,
        "headers": a.headers,
        "target": a.target,
        "entity_name": a.entity_name,
        "confidence": a.confidence,
        "column_mapping": a.column_mapping,
        "anomalies": a.anomalies,
        "rationale": a.rationale,
        "sample_rows": a.sample_rows,
        "row_count": a.row_count,
    }


def _persisted_table_payload(a: TableAnalysis, body_rows: list[list[Any]]) -> dict[str, Any]:
    """Persisted payload — adds the FULL bounded body (capped at MAX_BODY_ROWS)
    so /ai-sync can replay against real data, not just the LLM preview."""
    p = _public_table_payload(a)
    p["body_rows"] = [
        [to_jsonable(c) for c in r] for r in body_rows[:MAX_BODY_ROWS]
    ]
    return p


def _persist_analysis(
    client: Client, payload: AnalysisPersistPayload,
) -> str:
    aid = str(uuid.uuid4())
    bodies = {t.source: t.rows for t in payload.extracted}
    insert_payload = {
        "id": aid,
        "actor_user_id": payload.caller.get("user_id"),
        "actor_label": payload.caller.get("label"),
        "plant_id": payload.plant_id,
        "filename": payload.filename[:200],
        "file_kind": payload.file_kind[:16],
        "file_size": payload.file_size,
        "ai_provider": payload.ai_provider,
        "ai_model": payload.ai_model,
        "status": "pending",
        "wellmeter_detected": payload.wellmeter_detected,
        "tables": [
            _persisted_table_payload(a, bodies.get(a.source, []))
            for a in payload.analyses
        ],
    }
    try:
        client.table("import_analysis").insert(insert_payload).execute()
    except Exception as e:
        # Fail loudly — silent persistence failure was masking missing
        # migrations and producing false-positive analyse responses
        # whose IDs would later 404 on /ai-sync.
        raise HTTPException(
            status_code=500,
            detail=(
                f"Failed to persist analysis ({e}). "
                "Did you run supabase/migrations/20260425_import_analysis.sql?"
            ),
        )
    return aid


def _load_analysis(client: Client, analysis_id: str) -> dict[str, Any]:
    try:
        res = (
            client.table("import_analysis")
            .select("*").eq("id", analysis_id).maybe_single().execute()
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load analysis: {e}")
    if not res or not res.data:
        raise HTTPException(status_code=404, detail="Analysis not found.")
    return res.data


# ---------------------------------------------------------------------------
# Sync — entity creation + reading inserts
# ---------------------------------------------------------------------------
#
# The per-target row builders, column-resolver, date/float coercions, and
# the get-or-create entity helper all live in `ai_import_helpers.py` so
# this module can stay focused on Supabase orchestration + audit logging.


def _insert_readings(client: Client, ctx: ReadingsInsertContext) -> tuple[int, list[str]]:
    """Insert reading rows for `ctx.target`. Returns (count, warnings)."""
    payload, warnings = build_reading_rows(ctx)
    if not payload:
        return 0, warnings

    inserted = 0
    BATCH = 500
    for i in range(0, len(payload), BATCH):
        chunk = payload[i:i + BATCH]
        try:
            client.table(ctx.target).insert(chunk).execute()
            inserted += len(chunk)
        except Exception as e:  # noqa: BLE001
            warnings.append(f"Batch {i // BATCH + 1} failed: {e}")
            break
    return inserted, warnings


# ---------------------------------------------------------------------------
# Public entry points (called from server.py)
# ---------------------------------------------------------------------------

def analyze_upload(authorization: Optional[str], file: UploadFile, content: bytes,
                   plant_id: Optional[str]) -> dict[str, Any]:
    token = bearer_token(authorization)
    caller = caller_identity(token)
    require_roles(caller, {"Admin", "Manager"})
    client = user_scoped_client(token)

    filename = file.filename or "upload"
    tables = extract_tables(filename, content)
    if not tables:
        raise HTTPException(status_code=400, detail="No tables detected in the file.")

    analyses, provider = classify_tables(tables)
    wellmeter = looks_like_wellmeter(tables)
    file_kind = (filename.rsplit(".", 1)[-1] if "." in filename else "")[:16].lower()
    aid = _persist_analysis(
        client,
        AnalysisPersistPayload(
            caller=caller,
            filename=filename,
            file_kind=file_kind,
            file_size=len(content),
            plant_id=plant_id,
            extracted=tables,
            analyses=analyses,
            wellmeter_detected=wellmeter,
            ai_provider=provider,
            ai_model=OPENAI_MODEL if provider == "openai" else None,
        ),
    )

    return {
        "analysis_id": aid,
        "filename": filename,
        "wellmeter_detected": wellmeter,
        "ai_provider": provider,
        "ai_model": OPENAI_MODEL if provider == "openai" else None,
        "tables": [_public_table_payload(a) for a in analyses],
    }


def sync_analysis(authorization: Optional[str], analysis_id: str,
                  body: dict[str, Any]) -> dict[str, Any]:
    """Body shape:
        {
          "reason": "<min 5 chars>",
          "plant_id": "<uuid>",
          "decisions": [
            {
              "source": "<table source>",
              "action": "sync"|"reject",
              "target": "wells"|"locators"|"ro_trains"|"well_readings"|...,
              "entity_name": "Well 1",
              "column_mapping": {"date": "Date", "volume": "Volume"}
            },
            ...
          ]
        }
    """
    token = bearer_token(authorization)
    caller = caller_identity(token)
    require_roles(caller, {"Admin"})
    client = user_scoped_client(token)

    reason = (body.get("reason") or "").strip()
    if len(reason) < 5:
        raise HTTPException(status_code=400, detail="`reason` is required (min 5 chars).")
    plant_id = (body.get("plant_id") or "").strip() or None
    decisions = body.get("decisions") or []
    if not isinstance(decisions, list) or not decisions:
        raise HTTPException(status_code=400, detail="`decisions` must be a non-empty list.")

    analysis = _load_analysis(client, analysis_id)
    if analysis.get("status") != "pending":
        raise HTTPException(status_code=409, detail=f"Analysis already {analysis.get('status')}.")
    tables_by_source = {t["source"]: t for t in (analysis.get("tables") or [])}

    summary: dict[str, Any] = {
        "created": {"wells": 0, "locators": 0, "ro_trains": 0},
        "existing": {"wells": 0, "locators": 0, "ro_trains": 0},
        "inserted": {"well_readings": 0, "locator_readings": 0,
                     "ro_train_readings": 0, "power_readings": 0},
        "skipped": [],
        "rejected": [],
    }

    any_synced = False
    any_skipped = False

    for d in decisions:
        source = str(d.get("source") or "").strip()
        action = str(d.get("action") or "").strip().lower()
        if not source or source not in tables_by_source:
            summary["skipped"].append({"source": source, "reason": "unknown source"})
            any_skipped = True
            continue
        tbl = tables_by_source[source]
        target = str(d.get("target") or tbl.get("target") or "unknown")
        entity_name = (d.get("entity_name") or tbl.get("entity_name") or "").strip()
        mapping = dict(d.get("column_mapping") or tbl.get("column_mapping") or {})

        if action == "reject" or target in ("skip", "unknown"):
            summary["rejected"].append({"source": source, "target": target})
            _audit_decision(client, caller, AuditDecision(
                analysis_id=analysis_id, source=source, target=target,
                label="[IMPORT-REJECT]", reason=reason,
            ))
            continue

        if action != "sync":
            summary["skipped"].append({"source": source, "reason": f"unknown action '{action}'"})
            any_skipped = True
            continue

        if target in ENTITY_TARGETS:
            if not plant_id:
                summary["skipped"].append({"source": source, "reason": "plant_id required for entity creation"})
                any_skipped = True
                continue
            if not entity_name:
                summary["skipped"].append({"source": source, "reason": "entity_name required"})
                any_skipped = True
                continue
            ent_id, was_created = ensure_entity(client, target, plant_id, entity_name)
            if ent_id:
                if was_created:
                    summary["created"][target] += 1
                else:
                    summary["existing"][target] += 1
                any_synced = True
                tag = "[IMPORT]" if was_created else "[IMPORT-EXISTS]"
                _audit_decision(client, caller, AuditDecision(
                    analysis_id=analysis_id, source=source, target=target,
                    label=f"{tag} {entity_name}", reason=reason,
                ))
            else:
                summary["skipped"].append({"source": source, "reason": f"failed to create {target}"})
                any_skipped = True
            continue

        if target in READING_TARGETS:
            if not plant_id:
                summary["skipped"].append({"source": source, "reason": "plant_id required for readings"})
                any_skipped = True
                continue
            entity_id: Optional[str] = None
            if target != "power_readings":
                entity_table = {
                    "well_readings": "wells",
                    "locator_readings": "locators",
                    "ro_train_readings": "ro_trains",
                }[target]
                if not entity_name:
                    summary["skipped"].append({"source": source, "reason": "entity_name required for readings"})
                    any_skipped = True
                    continue
                entity_id, was_created = ensure_entity(client, entity_table, plant_id, entity_name)
                if not entity_id:
                    summary["skipped"].append({"source": source, "reason": f"could not resolve {entity_table} '{entity_name}'"})
                    any_skipped = True
                    continue
                if was_created:
                    summary["created"][entity_table] += 1
                else:
                    summary["existing"][entity_table] += 1

            # Use the FULL persisted body (capped at MAX_BODY_ROWS) — not the
            # sample rows. This makes /ai-sync actually import the data the
            # admin reviewed, rather than only the 25-row preview.
            body_rows = tbl.get("body_rows")
            if not body_rows:
                # Backwards-compat for analyses persisted before body_rows
                # existed: fall back to sample_rows so old `pending` rows
                # remain consumable.
                body_rows = tbl.get("sample_rows") or []
            inserted, warns = _insert_readings(
                client,
                ReadingsInsertContext(
                    target=target,
                    plant_id=plant_id,
                    entity_id=entity_id or "",
                    recorded_by=caller.get("user_id"),
                    headers=tbl.get("headers") or [],
                    rows=body_rows,
                    mapping=mapping,
                ),
            )
            summary["inserted"][target] += inserted
            if inserted:
                any_synced = True
                _audit_decision(client, caller, AuditDecision(
                    analysis_id=analysis_id, source=source, target=target,
                    label=f"[IMPORT] {entity_name or target} ({inserted} rows)",
                    reason=reason,
                ))
            else:
                # No rows landed — record this as a partial outcome with a
                # rejection-style audit row so it's traceable.
                _audit_decision(client, caller, AuditDecision(
                    analysis_id=analysis_id, source=source, target=target,
                    label="[IMPORT-REJECT] no readings imported", reason=reason,
                ))
            for w in warns:
                summary["skipped"].append({"source": source, "reason": w})
                any_skipped = True
            continue

        summary["skipped"].append({"source": source, "reason": f"unsupported target '{target}'"})
        any_skipped = True

    new_status = (
        "synced" if any_synced and not any_skipped
        else "partial" if any_synced
        else "rejected"
    )

    try:
        client.table("import_analysis").update({
            "status": new_status,
            "decisions": decisions,
            "reason": reason,
            "decided_by": caller.get("user_id"),
            "decided_at": datetime.now(timezone.utc).isoformat(),
            "sync_summary": summary,
        }).eq("id", analysis_id).execute()
    except Exception as e:
        # Surface the failure: a silent update means the analysis stays
        # `pending` forever and the admin has no way to know their decisions
        # weren't recorded. The downstream side-effects (entity rows + audit
        # log) already happened, so we report partial success in the error.
        log.exception("Failed to update import_analysis status")
        raise HTTPException(
            status_code=500,
            detail=(
                f"Sync side-effects ran but status update failed: {e}. "
                "Audit log + business rows reflect the actual state — "
                "but the analysis row is still 'pending'."
            ),
        )

    return {
        "ok": True,
        "analysis_id": analysis_id,
        "status": new_status,
        "summary": summary,
    }


def _audit_decision(
    client: Client, caller: dict[str, Any], decision: AuditDecision,
) -> None:
    """Write one row in deletion_audit_log so import decisions show up in the
    same admin Audit Log UI as deletions. We deliberately reuse the existing
    audit table (with kind='plant') and embed the [IMPORT] / [IMPORT-REJECT]
    tag in `reason` so existing filters (and the AuditLogPanel UI) keep
    working without a schema change."""
    write_audit(client, AuditEntry(
        kind="plant",
        entity_id=decision.analysis_id,
        entity_label=f"{decision.source} → {decision.target}",
        action="hard",
        caller=caller,
        reason=f"{decision.label} | {decision.reason}",
        dependencies={"source": decision.source, "target": decision.target},
    ))


def list_analyses(authorization: Optional[str], limit: int = 25) -> dict[str, Any]:
    token = bearer_token(authorization)
    caller = caller_identity(token)
    require_roles(caller, {"Admin", "Manager"})
    client = user_scoped_client(token)
    try:
        res = (
            client.table("import_analysis")
            .select("id,filename,file_kind,status,wellmeter_detected,ai_provider,ai_model,actor_label,plant_id,created_at,decided_at,sync_summary")
            .order("created_at", desc=True).limit(max(1, min(int(limit), 100))).execute()
        )
        return {"count": len(res.data or []), "entries": res.data or []}
    except Exception as e:
        log.warning("list_analyses failed: %s", e)
        return {"count": 0, "entries": [], "warning": str(e)}
