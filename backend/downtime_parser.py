"""
Parser for the 'Downtime' worksheet used by Mambaling etc.

Layout (per month block of 5 columns, repeated horizontally):
    Header row:   'Jan. 2026' | 'Total Operation, Hrs' | 'Total Shutdown, Hrs' | 'Remarks' | <blank>
    Data row:     <day>       | <op_hrs>                | <shutdown_hrs>        | <remarks> | ...

`Remarks` is free-form prose listing multiple events separated by commas/&,
for example:

    "Shutdown R.O #1:6hrs.&44mins. Due to Replacement of RO System., Shutdown
     Well #1:6hrs.&33mins. Due to High Raw Water Level."

This module splits a remark string into individual events, each with:
  - subsystem ('RO 1', 'Well 3', 'MCWD Supply', 'Well 1&4', …)
  - duration_hrs (float)
  - cause (trimmed string)
  - raw_text (original snippet)
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, asdict
from datetime import datetime, date
from typing import Any, Optional

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Regex building blocks
# ---------------------------------------------------------------------------

# "6hrs.&44mins."  or "24hrs."  or "45mins."  or "1hr.&10mins."
_DURATION_RE = re.compile(
    r"(?:(\d+(?:\.\d+)?)\s*(?:hrs?|hours?)\.?)"       # hours
    r"(?:\s*(?:\.\s*)?&\s*(\d+(?:\.\d+)?)\s*(?:mins?|minutes?)\.?)?"  # optional mins
    r"|"
    r"(\d+(?:\.\d+)?)\s*(?:mins?|minutes?)\.?",       # mins-only
    re.I,
)

# Prefix words that introduce an event
_EVENT_SPLIT_RE = re.compile(
    r"(?<![.\w])(?=Shutdown\b|Shut-?Down\b|\d+\s*hr)", re.I,
)

_SUBSYSTEM_RE = re.compile(
    r"\b("
    r"R\.?O\.?\s*#?\s*\d+(?:\s*(?:&|and)\s*#?\s*\d+)*"
    r"|Well\s*#?\s*\d+(?:\s*(?:&|and)\s*#?\s*\d+)*"
    r"|Well(?:'s)?"
    r"|MCWD(?:\s+Supply)?"
    r"|Product\s+Tank"
    r"|Booster\s+Pump"
    r"|System"
    r")\b",
    re.I,
)


def _normalize_sub(s: str) -> str:
    t = re.sub(r"\s+", " ", s).strip()
    t = re.sub(r"\bR\.?O\.?\b", "RO", t, flags=re.I)
    t = re.sub(r"\bWell'?s\b", "Wells", t, flags=re.I)
    t = re.sub(r"#\s*", "#", t)
    t = t.replace(" and ", " & ")
    return t


def _duration_hours(segment: str) -> float:
    """Sum up all hh/mm occurrences in the segment."""
    total = 0.0
    for m in _DURATION_RE.finditer(segment):
        hrs_str, mins_after_hrs, mins_only = m.groups()
        if hrs_str:
            total += float(hrs_str) + (float(mins_after_hrs) / 60.0 if mins_after_hrs else 0.0)
        elif mins_only:
            total += float(mins_only) / 60.0
    return round(total, 3)


def _clean_cause(raw: str) -> str:
    m = re.search(r"Due\s+to\s+(.+?)(?=\s*(?:Shutdown|$))", raw, flags=re.I)
    if not m:
        return ""
    cause = m.group(1).strip(" .,&")
    cause = re.sub(r"\s+", " ", cause)
    # Drop trailing fragment if it's just another event's duration ("6hrs.&44mins.")
    cause = re.sub(r"[,.]?\s*\d+\s*(?:hrs?|mins?)\.?.*$", "", cause, flags=re.I).strip()
    return cause[:240]


@dataclass
class DowntimeEvent:
    event_date: str        # ISO yyyy-mm-dd
    subsystem: str         # 'RO 1', 'Well 3', 'MCWD Supply' …
    duration_hrs: float
    cause: str
    raw_text: str
    op_hrs: Optional[float] = None
    shutdown_hrs: Optional[float] = None


# ---------------------------------------------------------------------------
# Remark splitter
# ---------------------------------------------------------------------------

def split_remarks(remark: str) -> list[str]:
    """Split a free-form remark into individual event strings."""
    if not remark:
        return []
    s = remark.strip()
    # Normalise separators: replace '., Shutdown' / ' &  Shutdown' with unified pipe
    s = re.sub(r"\.?\s*&\s*(?=Shutdown\b)", " | ", s, flags=re.I)
    s = re.sub(r"\.\s*,\s*(?=Shutdown\b)", " | ", s, flags=re.I)
    s = re.sub(r",\s*(?=Shutdown\b)", " | ", s, flags=re.I)
    s = re.sub(r"\.\s+(?=Shutdown\b)", " | ", s, flags=re.I)
    # Also split "hh. X…" type MCWD-first segments
    s = re.sub(r"\.\s+(?=\d+\s*hrs?\b)", " | ", s, flags=re.I)
    s = re.sub(r",\s*(?=\d+\s*hrs?\b)", " | ", s, flags=re.I)

    parts = [p.strip(" .,&") for p in s.split("|") if p.strip()]
    # Filter out "Normal Operation" / empty-ish
    parts = [p for p in parts if not re.fullmatch(r"normal\s+operation\.?", p, flags=re.I)]
    return parts


def parse_event(part: str, event_date: str,
                op_hrs: Optional[float],
                shutdown_hrs: Optional[float]) -> Optional[DowntimeEvent]:
    dur = _duration_hours(part)
    if dur <= 0:
        return None
    sub_m = _SUBSYSTEM_RE.search(part)
    sub = _normalize_sub(sub_m.group(0)) if sub_m else "Plant"
    # If no subsystem found but part starts with "X hrs. Downtime to Y …" → use Y.
    if not sub_m:
        m2 = re.search(r"downtime\s+to\s+([A-Z][A-Za-z\s]+?)(?:\s+Due|\.|,)", part, re.I)
        if m2:
            sub = _normalize_sub(m2.group(1))
    cause = _clean_cause(part)
    return DowntimeEvent(
        event_date=event_date,
        subsystem=sub,
        duration_hrs=dur,
        cause=cause,
        raw_text=part[:400],
        op_hrs=op_hrs,
        shutdown_hrs=shutdown_hrs,
    )


# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------

def _header_blocks(header: tuple[Any, ...]) -> list[tuple[int, date]]:
    """Find (col_index, first_of_month) for each month block in the header."""
    out: list[tuple[int, date]] = []
    for i, v in enumerate(header):
        if v is None:
            continue
        if isinstance(v, (datetime, date)):
            d = v.date() if isinstance(v, datetime) else v
            out.append((i, d.replace(day=1)))
        elif isinstance(v, str):
            m = re.match(
                r"^\s*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s*(\d{4})\s*$",
                v, re.I,
            )
            if m:
                month = datetime.strptime(m.group(1)[:3], "%b").month
                year = int(m.group(2))
                out.append((i, date(year, month, 1)))
    return out


def parse_downtime_sheet(ws) -> list[DowntimeEvent]:
    """Extract all downtime events from a 'Downtime' worksheet."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    header = all_rows[0]
    blocks = _header_blocks(header)
    if not blocks:
        return []

    events: list[DowntimeEvent] = []
    for row in all_rows[1:]:
        for col, anchor in blocks:
            day = row[col] if col < len(row) else None
            op_hrs = row[col + 1] if col + 1 < len(row) else None
            sd_hrs = row[col + 2] if col + 2 < len(row) else None
            remark = row[col + 3] if col + 3 < len(row) else None
            if day is None or remark is None:
                continue
            try:
                d = int(day)
            except (TypeError, ValueError):
                continue
            if not (1 <= d <= 31):
                continue
            try:
                iso = anchor.replace(day=d).isoformat()
            except ValueError:
                continue
            op = _to_float(op_hrs)
            sd = _to_float(sd_hrs)
            parts = split_remarks(str(remark))
            for p in parts:
                ev = parse_event(p, iso, op, sd)
                if ev:
                    events.append(ev)
    return events


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Public entry
# ---------------------------------------------------------------------------

def parse_downtime_xlsx(data: bytes, sheet_name: str = "Downtime") -> list[dict[str, Any]]:
    """Load workbook, locate the downtime sheet (case-insensitive match) and
    return a list of event dicts."""
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    target = None
    for sn in wb.sheetnames:
        if sn.strip().lower() == sheet_name.lower():
            target = wb[sn]
            break
    if target is None:
        return []
    return [asdict(ev) for ev in parse_downtime_sheet(target)]
