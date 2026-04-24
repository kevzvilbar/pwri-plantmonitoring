"""
XLSX Well/Meter reading importer.

Handles the tri-block monthly layout used by Mambaling / SRP / Umapad / Guizo plants:

    | Date | Initial | Final | Volume | Status | Date | Initial | Final | Volume | Status | ...

Auto-detects status markers such as:
  - "Defective Meter"           -> invalid (excluded from totals, flagged)
  - "No Operation"              -> downtime (volume=0, flagged)
  - "Shut-Off" / "Shutdown" /
    "Tripped-Off" / "Standby"   -> downtime
  - "Blend" / "Blend/Shutdown"  -> valid reading, tagged for audit
  - "New Meter Reading"         -> baseline reset
  - Any "No Reading Due to ..." -> downtime w/ reason

Also flags inconsistent rows (Final < Initial, duplicate dates).
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

# ---------------------------------------------------------------------------
# Status taxonomy
# ---------------------------------------------------------------------------

# Each raw marker maps to a normalized status code.
# Codes we produce:
#   valid            - normal reading, include in totals
#   blend            - valid but tagged (audit)
#   blend_shutdown   - partial blend, treat as valid w/ downtime partial
#   defective        - exclude from totals, flag
#   shutoff          - downtime, volume=0
#   no_operation     - downtime, volume=0
#   no_reading       - reading skipped (rain/manual/etc.) -> downtime-ish
#   new_meter        - baseline reset, still include
#   standby          - downtime
#   tripped          - downtime
#   unknown          - unknown raw label, flagged

STATUS_MAP: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\bdefective\b", re.I), "defective"),
    (re.compile(r"\bnew\s+meter\b", re.I), "new_meter"),
    (re.compile(r"\bblend\s*/\s*shut", re.I), "blend_shutdown"),
    (re.compile(r"\bblend\b", re.I), "blend"),
    (re.compile(r"\bshut[\s\-]?off\b", re.I), "shutoff"),
    (re.compile(r"\bshutdown\b", re.I), "shutoff"),
    (re.compile(r"\btripped[\s\-]?off\b", re.I), "tripped"),
    (re.compile(r"\bstandby\b", re.I), "standby"),
    (re.compile(r"\bno\s+operation\b", re.I), "no_operation"),
    (re.compile(r"\bno\s+reading\b", re.I), "no_reading"),
]

# Which statuses should count as downtime (volume forced to 0 for totals)
DOWNTIME_STATUSES = {
    "shutoff", "no_operation", "tripped", "standby", "no_reading", "blend_shutdown",
}

# Which statuses are excluded from totals
EXCLUDE_FROM_TOTALS = {"defective"}

# Which statuses are "valid" (count toward production)
VALID_STATUSES = {"valid", "blend", "new_meter"}


def classify_status(raw: Optional[str]) -> tuple[str, Optional[str]]:
    """Return (status_code, raw_trimmed_or_None)."""
    if raw is None:
        return "valid", None
    s = str(raw).strip()
    if not s:
        return "valid", None
    for pat, code in STATUS_MAP:
        if pat.search(s):
            return code, s
    return "unknown", s


# ---------------------------------------------------------------------------
# Row dataclass
# ---------------------------------------------------------------------------

@dataclass
class ParsedRow:
    date: Optional[str]  # ISO yyyy-mm-dd
    initial: Optional[float]
    final: Optional[float]
    volume: Optional[float]
    status: str                     # normalized status code
    status_raw: Optional[str]       # raw label
    include_in_totals: bool
    is_downtime: bool
    flags: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    row_index: int = 0              # source row in xlsx
    block_index: int = 0            # which month-block (0,1,2,...)


@dataclass
class SheetResult:
    sheet_name: str
    suggested_well_name: str
    rows: list[ParsedRow]
    summary: dict[str, Any]
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _as_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _as_date(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    # Try parse string
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _clean_sheet_name(name: str) -> str:
    """Extract a short well name e.g. 'Well 2 Meter Reading' -> 'Well 2'."""
    s = name
    s = re.sub(r"meter\s+reading", "", s, flags=re.I)
    s = re.sub(r"\(.*?\)", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s or name


# ---------------------------------------------------------------------------
# Core parser
# ---------------------------------------------------------------------------

def _detect_blocks(header_row: tuple[Any, ...]) -> list[tuple[int, Optional[date]]]:
    """
    Given the first row cells, find column indices (0-based) where each month
    block starts. A block starts at a column whose value is a date/datetime or
    the literal string 'Date'.

    Returns list of (col_index, anchor_month_start_or_None).
    The anchor is used when data rows contain only a day-of-month integer.
    """
    starts: list[tuple[int, Optional[date]]] = []
    for i, val in enumerate(header_row):
        if val is None:
            continue
        if isinstance(val, datetime):
            starts.append((i, val.date().replace(day=1)))
        elif isinstance(val, date):
            starts.append((i, val.replace(day=1)))
        elif isinstance(val, str) and val.strip().lower() == "date":
            starts.append((i, None))
    return starts


def _resolve_date(raw: Any, anchor: Optional[date]) -> Optional[str]:
    """
    Resolve a row's date cell into an ISO string.
    * If the cell is already a real datetime/date -> use it.
    * If it's an int/float in [1, 31] and we have an anchor month -> combine.
    * If it's a parseable date string -> parse.
    """
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        n = int(raw)
        if 1 <= n <= 31 and anchor is not None:
            try:
                return anchor.replace(day=n).isoformat()
            except ValueError:
                return None
        return None
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # Try pure integer day
    try:
        n = int(s)
        if 1 <= n <= 31 and anchor is not None:
            return anchor.replace(day=n).isoformat()
    except ValueError:
        pass
    return None


def _parse_sheet(ws) -> SheetResult:
    rows_out: list[ParsedRow] = []
    seen_dates: dict[str, int] = {}
    warnings_s: list[str] = []

    # use data_only values (formulas already evaluated and cached by Excel)
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return SheetResult(sheet_name=ws.title, suggested_well_name=_clean_sheet_name(ws.title),
                           rows=[], summary={"total_rows": 0}, warnings=["empty sheet"])

    header = all_rows[0]
    starts = _detect_blocks(header)
    if not starts:
        # fallback: assume single block starting at col 0
        starts = [(0, None)]

    # Iterate every data row (skip header r1)
    for r_idx, row in enumerate(all_rows[1:], start=2):
        for b_idx, (col, anchor) in enumerate(starts):
            # Each block is [Date, Initial, Final, Volume, Status?]
            date_v = row[col] if col < len(row) else None
            init_v = row[col + 1] if col + 1 < len(row) else None
            fin_v = row[col + 2] if col + 2 < len(row) else None
            vol_v = row[col + 3] if col + 3 < len(row) else None
            stat_v = row[col + 4] if col + 4 < len(row) else None

            # Skip completely empty row-block
            if (date_v in (None, "") and init_v in (None, "")
                    and fin_v in (None, "") and stat_v in (None, "")):
                continue

            # Many templates store only the day number in the "date" col.
            # We accept string dates, real datetimes, or None.
            iso = _resolve_date(date_v, anchor)
            # Fallback: look at the header row of this block for an anchor
            if iso is None and date_v is not None and anchor is None:
                iso = _as_date(date_v)

            initial = _as_float(init_v)
            final = _as_float(fin_v)
            volume = _as_float(vol_v)
            status_code, status_raw = classify_status(
                stat_v if isinstance(stat_v, str) else None
            )

            flags: list[str] = []
            warnings: list[str] = []

            # Compute volume if missing and both init+fin available
            if volume is None and initial is not None and final is not None:
                volume = final - initial

            # Inconsistent: Final < Initial
            if initial is not None and final is not None and final < initial:
                flags.append("inconsistent")
                warnings.append(
                    f"Final ({final}) < Initial ({initial})")

            # Duplicate-date detection (per sheet)
            if iso:
                if iso in seen_dates:
                    flags.append("duplicate_date")
                    warnings.append(f"Duplicate of row {seen_dates[iso]}")
                else:
                    seen_dates[iso] = r_idx

            # Decide include/exclude & downtime
            is_downtime = status_code in DOWNTIME_STATUSES
            include = status_code not in EXCLUDE_FROM_TOTALS and "inconsistent" not in flags

            # For downtime rows, force volume to 0 for totals (keep raw for display)
            if is_downtime and (volume is None or (isinstance(volume, float) and volume < 0)):
                volume = 0

            if status_code == "unknown" and status_raw:
                flags.append("unknown_status")
                warnings.append(f"Unrecognized status '{status_raw}'")

            rows_out.append(ParsedRow(
                date=iso,
                initial=initial,
                final=final,
                volume=volume,
                status=status_code,
                status_raw=status_raw,
                include_in_totals=include,
                is_downtime=is_downtime,
                flags=flags,
                warnings=warnings,
                row_index=r_idx,
                block_index=b_idx,
            ))

    # Sort rows by date ascending (None at end)
    rows_out.sort(key=lambda r: (r.date is None, r.date or "", r.block_index))

    # Summary
    by_status: dict[str, int] = {}
    for r in rows_out:
        by_status[r.status] = by_status.get(r.status, 0) + 1

    sum_volume_valid = sum(
        (r.volume or 0) for r in rows_out
        if r.include_in_totals and not r.is_downtime and r.status in VALID_STATUSES
    )
    sum_volume_all = sum((r.volume or 0) for r in rows_out if r.include_in_totals)

    flagged = sum(1 for r in rows_out if r.flags)
    defective = by_status.get("defective", 0)
    downtime = sum(1 for r in rows_out if r.is_downtime)

    summary = {
        "total_rows": len(rows_out),
        "by_status": by_status,
        "valid_rows": sum(1 for r in rows_out if r.include_in_totals and not r.is_downtime),
        "defective_rows": defective,
        "downtime_rows": downtime,
        "flagged_rows": flagged,
        "sum_volume_valid": round(sum_volume_valid, 3),
        "sum_volume_all_included": round(sum_volume_all, 3),
        "date_range": [
            next((r.date for r in rows_out if r.date), None),
            next((r.date for r in reversed(rows_out) if r.date), None),
        ],
    }

    return SheetResult(
        sheet_name=ws.title,
        suggested_well_name=_clean_sheet_name(ws.title),
        rows=rows_out,
        summary=summary,
        warnings=warnings_s,
    )


def parse_xlsx(data: bytes) -> dict[str, Any]:
    """
    Main entry. Returns a dict suitable for JSON response.
    """
    bio = io.BytesIO(data)
    # data_only=True reads the cached computed values Excel stored, which
    # is what we want since many cells are formulas like =C2-B2.
    wb = load_workbook(bio, data_only=True, read_only=True)

    sheets: list[dict[str, Any]] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        # Skip sheets with too few columns/rows to be a meter sheet
        if ws.max_row is None or ws.max_column is None:
            continue
        if ws.max_row < 2:
            continue
        result = _parse_sheet(ws)
        sheets.append({
            "sheet_name": result.sheet_name,
            "suggested_well_name": result.suggested_well_name,
            "rows": [asdict(r) for r in result.rows],
            "summary": result.summary,
            "warnings": result.warnings,
        })

    # File-level summary
    total_rows = sum(s["summary"]["total_rows"] for s in sheets)
    total_defective = sum(s["summary"]["defective_rows"] for s in sheets)
    total_downtime = sum(s["summary"]["downtime_rows"] for s in sheets)
    total_flagged = sum(s["summary"]["flagged_rows"] for s in sheets)

    return {
        "sheets": sheets,
        "file_summary": {
            "sheet_count": len(sheets),
            "total_rows": total_rows,
            "total_defective": total_defective,
            "total_downtime": total_downtime,
            "total_flagged": total_flagged,
        },
    }
